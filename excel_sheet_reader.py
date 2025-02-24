import os
from typing import Optional
import pandas as pd
from openpyxl import load_workbook


class ExcelSheetReader:
    """Класс для чтения листов из Elsx файлов"""

    @staticmethod
    def read_excel_file(file_path: str, sheet_name: str,
                        skiprows: Optional[int] = None, nrows: Optional[int] = None,
                        usecols: Optional[list[int]] = None, specific_rows: Optional[list[int]] = None) -> pd.DataFrame:
        """Чтение таблицы из Excel файла"""
        # Формируем параметры считывания файла
        read_excel_params = {
            'sheet_name': sheet_name,  # Имя листа
            'engine': 'openpyxl',
            'skiprows': skiprows,  # Количество строк, которые нужно пропустить в начале файла.
            'usecols': usecols,  # Список столбцов, которые нужно считать.
            'nrows': nrows  # Количество строк, которые нужно считать после пропуска.
        }

        # Проверяем существование файла
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл {file_path} не найден")

        wb = load_workbook(filename=file_path, read_only=True)

        # Проверяем существование листа
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Лист {sheet_name} не найден в файле {file_path}.")

        df: pd.DataFrame = pd.read_excel(file_path, **read_excel_params)

        if specific_rows is not None:
            df = df.iloc[specific_rows]

        # Удаляем строки с пустыми значениями
        df.dropna(inplace=True)
        return df

